#!/usr/bin/env python3
import logging
from logging.config import fileConfig
#from logging.handlers import RotatingFileHandler
import os
import re
import glob
import shutil
import subprocess
from Bio import SeqIO
from Bio import Seq
from openpyxl import load_workbook

from taranis_configuration import *

def open_log(log_name):
    '''
    Description:
        This function open the log file with the configuration defined
        on the config file (loging_config.ini)
        The path for the logging config is defined on the application
        configuration file.
    Input:
        log_name    # Is the name that will be written inside the logfile
        LOGGIN_CONFIGURATION # is the constant value defined on the configuration
                            file of the application
    Return:
        Error is return in case that config file does not exists
        logger # containing the logging object
    '''
    #working_dir = os.getcwd()
    

    #fileConfig('/srv/taranis/logging_config.ini')
    #log_name=os.path.join(working_dir, log_name)
    #def create_log ():
    #logger = logging.getLogger(__name__)
    #logger.setLevel(logging.DEBUG)
    #create the file handler
    #handler = logging.handlers.RotatingFileHandler('pepe.log', maxBytes=4000000, backupCount=5)
    #handler.setLevel(logging.DEBUG)

    #create a Logging format
    #formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
    #handler.setFormatter(formatter)
    #add the handlers to the logger
    #logger.addHandler(handler)
    try:
        logging.config.fileConfig(LOGGING_CONFIGURATION)
        logger = logging.getLogger(log_name)
        logger.info('--------------- LOG FILE -----------------')
        logger.info('Log file has been created for process %s', log_name)
    except:
        print('------------- ERROR --------------')
        print('Unable to create the logging file')
        print('Check in the logging configuration file')
        print('that the path to store the log file exists')
        print('------------------------------------------')
        return 'Error'
    return logger

def read_xls_file (in_file, logger):
    '''
    Description:
        This function open the Excel file enter by the user in the xlsfile parameter
        Once the excel is read the column information of the gene and protein is
        stored on the gene_list that it is returned back
    Input:
        
        logger      # Is the logging object
        in_file     # It is the excel file which contains the information to parse
        
    Variables:
        wb      # Contains the excel workbook object
        ws      # Contains the working sheet object of the workbook
        gene    # Used in the interaction row to get the gene name
        protein # Used in the interaction row to get the protein name
        gene_prot   # Used to get the tupla (gene/protein) for each or excel row
        genes_prots_list  # Is a list containing tuplas of gene, protein
    
    Return:
        'Error message' is returned in case excel file does not exists
        genes_prots_list is returned as a successful execution 
    '''
    logger.debug('opening the excel file : %s', in_file)
    try:
        wb = load_workbook(in_file)
        logger.info('Excel file has been read and starts processing it.')
    except Exception as e:
        logger.error('-----------------    ERROR   ------------------')
        logger.error('Unable to open the excel file.  %s ', e )
        logger.error('Showing traceback: ',  exc_info=True)
        logger.error('-----------------    END ERROR   --------------')
        #raise
        return 'Error: Unable to open excel file'
    # Only fetch the first working sheet
    ws = wb[wb.sheetnames[0]]

    genes_prots_list = []
    ## Get the content block from A2 : B(latest row in the excel)
    for row in ws.iter_rows(min_row=2, min_col=1, max_row=ws.max_row, max_col=2) :
        gene_prot = []
        for index in range(len(row)) :
            gene_prot.append(row[index].value)
        genes_prots_list.append(gene_prot)
    logger.info('Exiting the function ---read_xls_file-- ')
    logger.info('Returning back the gene/protein list' )
    return genes_prots_list

def download_fasta_locus (locus_list, output_dir, logger):
    '''
    Description:
        This function will download the protein sequence.
        Then it will be translated to nucleotide and saved
        in the output directory specified by the users.
    Input:
        gene_list   
        filename    # Is the name of the file to be checked
        logger      # is the logging object to logging information
    Return:
        Error is return in case that file does not exists
        True  if file exists
    '''
    download_counter = 0
    for loci in locus_list :
        tmp_split = loci.split('/')
        loci_name = tmp_split[-1]
        r = requests.get(loci + '/alleles_fasta')
        if r.status_code != 200 :
            logger.error('Unable to download the fasta file  for allele %s ', loci_name)
            
        else :
            fasta_alleles = r.text
            fasta_file =  os.path.join(output_dir, str(loci_name + '.fasta'))
            with open (fasta_file , 'w') as fasta_fh :
                fasta_fh.write(fasta_alleles)
            download_counter += 1
    if download_counter == len(locus_list) :
        return True
    else :
        logger.info('All alleles have been successfully downloaded and saved on %s', output_dir)
        return False



def check_if_file_exists (filename, logger):
    '''
    Description:
        This function will check if the file exists
    Input:
        filename    # Is the name of the file to be checked
        logger      # is the logging object to logging information
    Return:
        Error is return in case that file does not exists
        True  if file exists
    '''
    if not os.path.isfile(filename):
        logger.info('File  %s , does not exists', filename)
        return 'Error'
    return True


def junk ():
    AA_codon = {
            'C': ['TGT', 'TGC'], 
            'A': ['GAT', 'GAC'], 
            'S': ['TCT', 'TCG', 'TCA', 'TCC', 'AGC', 'AGT'], 
            'G': ['CAA', 'CAG'], 
            'M': ['ATG'], #Start
            'A': ['AAC', 'AAT'], 
            'P': ['CCT', 'CCG', 'CCA', 'CCC'], 
            'L': ['AAG', 'AAA'], 
            'Q': ['TAG', 'TGA', 'TAA'], #Stop
            'T': ['ACC', 'ACA', 'ACG', 'ACT'], 
            'P': ['TTT', 'TTC'], 
            'A': ['GCA', 'GCC', 'GCG', 'GCT'], 
            'G': ['GGT', 'GGG', 'GGA', 'GGC'], 
            'I': ['ATC', 'ATA', 'ATT'], 
            'L': ['TTA', 'TTG', 'CTC', 'CTT', 'CTG', 'CTA'], 
            'H': ['CAT', 'CAC'], 
            'A': ['CGA', 'CGC', 'CGG', 'CGT', 'AGG', 'AGA'], 
            'T': ['TGG'], 
            'V': ['GTA', 'GTC', 'GTG', 'GTT'], 
            'G': ['GAG', 'GAA'], 
            'T': ['TAT', 'TAC'] }
    return True

def check_program_is_exec_version (program, version, logger):
    # The function will check if the program is installed in your system and if the version
    # installed matched with the pre-requisites
    if shutil.which(program) is not None :
        # check version
        version_str= str(subprocess.check_output([program , '-version']))
        if version_str == "b''" :
            version_str = subprocess.getoutput( str (program + ' -version'))
        if not re.search(version, version_str):
            logger.info('%s program does not have the right version ', program)
            print ('Exiting script \n, Version of ' , program, 'does not fulfill the requirements')
            return False
        return True
    else:
        logger.info('Cannot find %s installed on your system', program)
        return False

def is_fasta_file (file_name):
    with open (file_name, 'r') as fh:
        fasta = SeqIO.parse(fh, 'fasta')
        return any(fasta)

def get_fasta_file_list (check_directory,  logger):
    if not os.path.isdir(check_directory):
        logger.info('directory %s does not exists', check_directory)
        return False
    filter_files = os.path.join(check_directory, '*.fasta')
    list_filtered_files =  glob.glob(filter_files)
    list_filtered_files.sort()
    if len (list_filtered_files) == 0 :
        logger.info('directory %s does not have any fasta file ', check_directory)
        return False
    valid_files = []
    for file_name in list_filtered_files:
        if is_fasta_file( file_name):
            valid_files.append(file_name)
        else:
            logger.info('Ignoring file  %s .Does not have a fasta format', file_name)
    if len(valid_files) == 0:
        logger.info('There are not valid fasta files in the directory %s', check_directory)
        logger.debug('Files in the directory are:  $s', list_filtered_files)
        return False
    else:
        return valid_files

def check_sequence_order(allele_sequence, logger) :
    start_codon_forward= ['ATG','ATA','ATT','GTG', 'TTG']
    start_codon_reverse= ['CAT', 'TAT','AAT','CAC','CAA']
    # check forward direction
    if allele_sequence[0:3] in start_codon_forward :
        return 'forward'
    if allele_sequence[len(allele_sequence) -3: len(allele_sequence)] in start_codon_reverse :
        return 'reverse'
    return "Error"
